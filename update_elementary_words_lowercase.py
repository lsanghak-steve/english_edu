import os
import sys
import openpyxl
from supabase import create_client

sys.stdout.reconfigure(encoding='utf-8')

supabase_url = 'https://sqonhhqosyszncjfoxfd.supabase.co'
supabase_key = 'sb_publishable_1trPlZQEdVKMvUYQNV5aVA_nSQqOiuo'
supabase = create_client(supabase_url, supabase_key)

def update_excel_elementary_to_lowercase():
    excel_path = os.path.join(os.path.dirname(__file__), 'word', 'elementary_words_ko_zh.xlsx')
    if not os.path.exists(excel_path):
        print(f"⚠️ 엑셀 파일 없음: {excel_path}")
        return

    print("====================================================")
    print("📖 [2단계: 초등 단어장 엑셀 파일 단어 소문자 일괄 변환]")
    print("====================================================\n")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 헤더 확인
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    word_col_idx = 2  # 기본 2번째 열(B열)

    for idx, h in enumerate(headers, 1):
        if h in ['word', '단어', 'english']:
            word_col_idx = idx
            break

    count = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=word_col_idx)
        val = str(cell.value or '').strip()
        if val and val != val.lower():
            cell.value = val.lower()
            count += 1

    wb.save(excel_path)
    print(f"🎉 엑셀 초등단어장 총 {count}개 단어를 소문자로 변환 저장 완료!\n")

def update_supabase_elementary_to_lowercase():
    print("====================================================")
    print("🗄️ [3단계: Supabase 클라우드 DB 초등 단어 소문자 일괄 갱신]")
    print("====================================================\n")

    res = supabase.from_('words').select('id, word, grade_level, category, image_url').execute()
    words = res.data or []

    print(f"📋 총 {len(words)}개 단어 조회 완료")

    update_count = 0
    for w in words:
        raw_word = str(w.get('word') or '').strip()
        grade = w.get('grade_level') or w.get('category') or ''

        # 초등단어이거나 모든 단어를 소문자로 안전 변환
        lower_word = raw_word.lower()
        new_img_url = f"https://sqonhhqosyszncjfoxfd.supabase.co/storage/v1/object/public/word_images/{lower_word}.png"

        if raw_word != lower_word or w.get('image_url') != new_img_url:
            update_data = {
                'word': lower_word,
                'image_url': new_img_url
            }
            supabase.from_('words').update(update_data).eq('id', w['id']).execute()
            update_count += 1
            if update_count <= 10 or update_count % 100 == 0:
                print(f"  ✅ [{update_count}] ID:{w['id']} {raw_word} ➔ {lower_word}")

    print(f"\n🎉 Supabase DB 총 {update_count}개 단어 및 image_url 소문자 일괄 갱신 완료!\n")

if __name__ == '__main__':
    update_excel_elementary_to_lowercase()
    update_supabase_elementary_to_lowercase()
