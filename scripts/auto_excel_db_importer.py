import openpyxl, glob, json, sys, io, os
import urllib.request

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SUPABASE_URL = 'https://sqonhhqosyszncjfoxfd.supabase.co'
SUPABASE_KEY = 'sb_publishable_1trPlZQEdVKMvUYQNV5aVA_nSQqOiuo'

def fetch_all_db_words():
    """Supabase DB에서 기존에 등록된 전체 단어 목록을 로드합니다."""
    all_rows = []
    offset = 0
    limit = 1000
    while True:
        req = urllib.request.Request(
            f'{SUPABASE_URL}/rest/v1/words?select=id,word,phonics,meaning,category,grade_level,example_en,example_ko,image_url&offset={offset}&limit={limit}',
            headers={
                'apikey': SUPABASE_KEY,
                'Authorization': f'Bearer {SUPABASE_KEY}'
            }
        )
        try:
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read().decode('utf-8'))
                if not data:
                    break
                all_rows.extend(data)
                if len(data) < limit:
                    break
                offset += limit
        except Exception as e:
            print('❌ DB 단어 로드 중 오류:', e)
            break
    return all_rows

def get_header_index(headers, possible_names):
    """헤더 리스트에서 가능한 이름 중 첫 번째 일치 항목의 인덱스(1-based)를 반환합니다."""
    for idx, h in enumerate(headers, 1):
        if not h:
            continue
        h_str = str(h).strip()
        for name in possible_names:
            if name.lower() in h_str.lower():
                return idx
    return None

def process_excel_files():
    print('===========================================================')
    print('🚀 [단어 엑셀 ➔ DB 자동 매칭 & 신규 단어 자동 추가 엔진 v1.0]')
    print('===========================================================\n')

    # 1. DB 단어 로드
    db_words = fetch_all_db_words()
    print(f'📊 Supabase DB 현재 총 어휘 수: {len(db_words)}개')

    db_map = {}  # key: (grade, word.lower()) -> db_row
    max_id = 0
    for row in db_words:
        w_id = int(row['id'])
        if w_id > max_id:
            max_id = w_id
        w_str = (row.get('word') or '').strip().lower()
        grade = (row.get('grade_level') or '').strip()
        db_map[(grade, w_str)] = row
        if w_str not in db_map:
            db_map[w_str] = row

    # 2. word/ 및 word/multilingual/ 폴더 내 모든 xlsx 파일 스캔
    excel_files = glob.glob('word/**/*.xlsx', recursive=True) + glob.glob('word/*.xlsx')
    excel_files = sorted(list(set(excel_files)))

    print(f'🔍 총 {len(excel_files)}개의 엑셀 파일 감지:')
    for ef in excel_files:
        print(f'  - {ef}')
    print('')

    parsed_words = {} # key: (grade, word_lower) -> word_dict

    for ef in excel_files:
        try:
            wb = openpyxl.load_workbook(ef, data_only=True)
            for sname in wb.sheetnames:
                ws = wb[sname]
                if ws.max_row < 2:
                    continue

                headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]

                col_word = get_header_index(headers, ['영어 단어', '영어 (Word)', 'Word'])
                if not col_word:
                    continue

                col_phonics = get_header_index(headers, ['발음기호', 'IPA', 'phonics'])
                col_meaning_ko = get_header_index(headers, ['한국어 뜻', 'Meaning (KO)', 'Meaning'])
                col_meaning_zh = get_header_index(headers, ['中文释义', 'Meaning ZH-CN', 'Meaning ZH'])
                col_meaning_fr = get_header_index(headers, ['Français (Meaning FR)', 'Meaning FR', 'Français'])
                col_category = get_header_index(headers, ['주제', 'Category'])
                col_grade = get_header_index(headers, ['단계', 'Level', 'Grade'])
                col_example_en = get_header_index(headers, ['영어 예문', 'Example EN'])
                col_example_ko = get_header_index(headers, ['한국어 해석', 'Example KO'])
                col_example_zh = get_header_index(headers, ['中文例句', 'Example ZH-CN', 'Example ZH'])
                col_example_fr = get_header_index(headers, ['Phrase française', 'Example FR'])

                # 파일명/시트명 기반 Grade 결정
                file_grade = '초등단어'
                if 'middle' in ef.lower() or '중등' in ef.lower():
                    file_grade = '중등단어'
                elif 'highschool' in ef.lower() or '고등' in ef.lower() or 'suneung' in ef.lower():
                    file_grade = '고등/수능'

                for r in range(2, ws.max_row+1):
                    w_val = ws.cell(r, col_word).value if col_word else None
                    if not w_val:
                        continue
                    w_str = str(w_val).strip()
                    if not w_str or w_str.lower() in ('none', 'word', '영어 단어'):
                        continue

                    phonics = str(ws.cell(r, col_phonics).value or '').strip() if col_phonics else ''
                    m_ko = str(ws.cell(r, col_meaning_ko).value or '').strip() if col_meaning_ko else ''
                    m_zh = str(ws.cell(r, col_meaning_zh).value or '').strip() if col_meaning_zh else ''
                    m_fr = str(ws.cell(r, col_meaning_fr).value or '').strip() if col_meaning_fr else ''
                    cat = str(ws.cell(r, col_category).value or '').strip() if col_category else ''
                    grd = str(ws.cell(r, col_grade).value or '').strip() if col_grade else file_grade
                    if not grd or grd == 'None':
                        grd = file_grade

                    ex_en = str(ws.cell(r, col_example_en).value or '').strip() if col_example_en else ''
                    ex_ko = str(ws.cell(r, col_example_ko).value or '').strip() if col_example_ko else ''
                    ex_zh = str(ws.cell(r, col_example_zh).value or '').strip() if col_example_zh else ''
                    ex_fr = str(ws.cell(r, col_example_fr).value or '').strip() if col_example_fr else ''

                    key = (grd, w_str.lower())
                    if key not in parsed_words:
                        parsed_words[key] = {
                            'word': w_str,
                            'phonics': phonics if phonics != 'None' else '',
                            'meaning': m_ko if m_ko != 'None' else '',
                            'meaning_zh': m_zh if m_zh != 'None' else '',
                            'meaning_fr': m_fr if m_fr != 'None' else '',
                            'category': cat if cat != 'None' else '',
                            'grade_level': grd,
                            'example_en': ex_en if ex_en != 'None' else '',
                            'example_ko': ex_ko if ex_ko != 'None' else '',
                            'example_zh': ex_zh if ex_zh != 'None' else '',
                            'example_fr': ex_fr if ex_fr != 'None' else ''
                        }
                    else:
                        # 누락된 의미/예문 보충
                        item = parsed_words[key]
                        if not item['meaning_zh'] and m_zh and m_zh != 'None': item['meaning_zh'] = m_zh
                        if not item['meaning_fr'] and m_fr and m_fr != 'None': item['meaning_fr'] = m_fr
                        if not item['example_zh'] and ex_zh and ex_zh != 'None': item['example_zh'] = ex_zh
                        if not item['example_fr'] and ex_fr and ex_fr != 'None': item['example_fr'] = ex_fr

        except Exception as e:
            print(f'⚠️ 파일 처리 중 예외 발생 ({ef}):', e)

    print(f'✅ 엑셀 스캔 결과 총 {len(parsed_words)}개 고유 단어 추출 완료!')

    # 3. DB 비교 ➔ 신규 추가 리스트 & 기존 업데이트 리스트 분리
    new_words_to_insert = []
    existing_words_to_update = []

    next_new_id = max_id + 1

    for (grd, w_lower), item in parsed_words.items():
        db_match = db_map.get((grd, w_lower)) or db_map.get(w_lower)

        if not db_match:
            # 🆕 DB에 없는 완전 신규 단어 ➔ 자동 추가!
            item['id'] = next_new_id
            item['image_url'] = f"https://sqonhhqosyszncjfoxfd.supabase.co/storage/v1/object/public/word_images/{item['word'].capitalize()}.png"
            new_words_to_insert.append(item)
            next_new_id += 1
        else:
            # 🔄 기존 DB 단어 ➔ 프랑스어 / 중국어 뜻 & 예문 업데이트
            upd_item = {
                'id': db_match['id'],
                'word': db_match['word']
            }
            has_update = False
            if item['meaning_zh']:
                upd_item['meaning_zh'] = item['meaning_zh']
                has_update = True
            if item['example_zh']:
                upd_item['example_zh'] = item['example_zh']
                has_update = True
            if item['meaning_fr']:
                upd_item['meaning_fr'] = item['meaning_fr']
                has_update = True
            if item['example_fr']:
                upd_item['example_fr'] = item['example_fr']
                has_update = True

            if has_update:
                existing_words_to_update.append(upd_item)

    print(f'📌 [검증 결과]')
    print(f'  - 🆕 DB 신규 추가할 단어: {len(new_words_to_insert)}개')
    print(f'  - 🔄 다국어(프랑스어/중국어) 동기화 업데이트할 단어: {len(existing_words_to_update)}개\n')

    # 4. Supabase DB Upsert (새 단어 + 기존 단어 다국어 정보)
    all_db_payloads = new_words_to_insert + existing_words_to_update

    if all_db_payloads:
        chunk_size = 100
        success_count = 0
        for i in range(0, len(all_db_payloads), chunk_size):
            chunk = all_db_payloads[i:i+chunk_size]
            payload_bytes = json.dumps(chunk, ensure_ascii=False).encode('utf-8')

            req = urllib.request.Request(
                f'{SUPABASE_URL}/rest/v1/words',
                data=payload_bytes,
                headers={
                    'apikey': SUPABASE_KEY,
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'Content-Type': 'application/json',
                    'Prefer': 'resolution=merge-duplicates'
                },
                method='POST'
            )
            try:
                with urllib.request.urlopen(req) as resp:
                    if resp.status in (200, 201, 204):
                        success_count += len(chunk)
            except urllib.error.HTTPError as e:
                err_msg = e.read().decode('utf-8')
                print(f'⚠️ DB 전송 실패 (Chunk {i}):', err_msg)
                if 'meaning_fr' in err_msg or 'PGRST204' in err_msg:
                    print('📌 SQL 스크립트 실행 필요: sql/add_french_columns_to_words.sql 구문을 Supabase SQL Editor에 실행해주세요.')

        print(f'🎉 DB 업로드 완수: 총 {success_count}개 단어 DB 저장 완료!')

    # 5. 로컬 JSON 데이터셋 업데이트 (data/ 및 next_app/data/)
    local_json_files = [
        'data/parsed_800_words.json',
        'data/middle_school_words.json',
        'data/highschool_parsed.json'
    ]

    for ljf in local_json_files:
        if os.path.exists(ljf):
            with open(ljf, 'r', encoding='utf-8-sig') as f:
                content = json.load(f)

            upd_cnt = 0
            for row in content:
                w_str = (row.get('word') or '').strip().lower()
                # parsed_words에서 조회
                found = None
                for (grd, w_k), p_item in parsed_words.items():
                    if w_k == w_str:
                        found = p_item
                        break

                if found:
                    if found.get('meaning_fr'):
                        row['meaning_fr'] = found['meaning_fr']
                        row['meaningFr'] = found['meaning_fr']
                    if found.get('example_fr'):
                        row['example_fr'] = found['example_fr']
                        row['exampleFr'] = found['example_fr']
                    if found.get('meaning_zh'):
                        row['meaning_zh'] = found['meaning_zh']
                        row['meaningZh'] = found['meaning_zh']
                    if found.get('example_zh'):
                        row['example_zh'] = found['example_zh']
                        row['exampleZh'] = found['example_zh']
                    upd_cnt += 1

            with open(ljf, 'w', encoding='utf-8') as f:
                json.dump(content, f, ensure_ascii=False, indent=2)

            print(f'📝 로컬 {ljf} {upd_cnt}개 단어 프랑스어/중국어 다국어 동기화 완료!')

    print('\n===========================================================')
    print('✨ [완료] 단어 엑셀 DB 자동 검증 & 프랑스어/다국어 동기화 완수!')
    print('===========================================================')

if __name__ == '__main__':
    process_excel_files()
