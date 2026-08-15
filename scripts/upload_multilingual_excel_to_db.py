import openpyxl, os, sys, io, json, time, urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SUPABASE_URL = 'https://sqonhhqosyszncjfoxfd.supabase.co'
SUPABASE_KEY = 'sb_publishable_1trPlZQEdVKMvUYQNV5aVA_nSQqOiuo'

def clean_val(val):
    if val is None:
        return ''
    val_str = str(val).strip()
    return val_str

def parse_elementary():
    fp = r'D:\Yc_Back_Live\영창정밀_노트북\english_edu\word\multilingual\elementary_words_multilingual.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    hmap = {h: idx for idx, h in enumerate(headers)}

    items = []
    for r in rows[1:]:
        num = r[hmap.get('번호', 0)]
        if not num: continue
        db_id = int(num)
        word = clean_val(r[hmap.get('영어 단어 (Word)', 1)])
        phonics = clean_val(r[hmap.get('발음기호 (IPA)', 2)])
        meaning = clean_val(r[hmap.get('한국어 뜻 (Meaning)', 3)])
        meaning_zh = clean_val(r[hmap.get('中文释义 (Meaning ZH-CN)', 4)])
        meaning_fr = clean_val(r[hmap.get('Français (Meaning FR)', 5)])
        meaning_ja = clean_val(r[hmap.get('日本語の意味 (Meaning JA)', 6)])
        meaning_hi = clean_val(r[hmap.get('हिन्दी अर्थ (Meaning HI)', 7)])
        meaning_vi = clean_val(r[hmap.get('Nghĩa tiếng Việt (Meaning VI)', 8)])
        category = clean_val(r[hmap.get('주제 (Category)', 9)])
        ex_en = clean_val(r[hmap.get('영어 예문 (Example EN)', 12)])
        ex_ko = clean_val(r[hmap.get('한국어 해석 (Example KO)', 13)])
        ex_zh = clean_val(r[hmap.get('中文例句 (Example ZH-CN)', 14)])
        ex_fr = clean_val(r[hmap.get('Phrase française (Example FR)', 15)])
        ex_ja = clean_val(r[hmap.get('日本語の例文 (Example JA)', 16)])
        ex_hi = clean_val(r[hmap.get('हिन्दी उदाहरण (Example HI)', 17)])
        ex_vi = clean_val(r[hmap.get('Câu ví dụ tiếng Việt (Example VI)', 18)])

        item = {
            'id': db_id,
            'word': word,
            'phonics': phonics,
            'meaning': meaning,
            'meaning_zh': meaning_zh,
            'meaning_fr': meaning_fr,
            'meaning_ja': meaning_ja,
            'meaning_hi': meaning_hi,
            'meaning_vi': meaning_vi,
            'category': category,
            'grade_level': '초등단어',
            'grade_level_ko': '초등단어',
            'grade_level_zh': '小学英语',
            'grade_level_fr': 'Anglais Primaire',
            'grade_level_ja': '小学生英語',
            'grade_level_vi': 'Tiếng Anh Tiểu học',
            'grade_level_hi': 'प्राथमिक अंग्रेजी',
            'example_en': ex_en,
            'example_ko': ex_ko,
            'example_zh': ex_zh,
            'example_fr': ex_fr,
            'example_ja': ex_ja,
            'example_hi': ex_hi,
            'example_vi': ex_vi,
        }
        items.append(item)
    print(f"Parsed {len(items)} elementary words.")
    return items

def parse_middle():
    fp = r'D:\Yc_Back_Live\영창정밀_노트북\english_edu\word\multilingual\middle_school_words_multilingual.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    hmap = {h: idx for idx, h in enumerate(headers)}

    items = []
    for r in rows[1:]:
        num = r[hmap.get('번호', 0)]
        if not num: continue
        db_id = int(num) + 1000 # 1~1200 ➔ 1001~2200
        word = clean_val(r[hmap.get('영어 (Word)', 1)])
        phonics = clean_val(r[hmap.get('발음기호 (IPA)', 2)])
        meaning = clean_val(r[hmap.get('한국어 뜻 (Meaning)', 4)])
        meaning_zh = clean_val(r[hmap.get('中文释义 (Meaning ZH-CN)', 5)])
        meaning_fr = clean_val(r[hmap.get('Français (Meaning FR)', 6)])
        meaning_ja = clean_val(r[hmap.get('日本語の意味 (Meaning JA)', 7)])
        meaning_hi = clean_val(r[hmap.get('हिन्दी अर्थ (Meaning HI)', 8)])
        meaning_vi = clean_val(r[hmap.get('Nghĩa tiếng Việt (Meaning VI)', 9)])
        category_raw = clean_val(r[hmap.get('주제 (Category)', 10)])
        category = f"중등 - {category_raw}" if category_raw and not category_raw.startswith('중등') else category_raw
        ex_en = clean_val(r[hmap.get('영어 예문 (Example EN)', 11)])
        ex_ko = clean_val(r[hmap.get('한국어 해석 (Example KO)', 12)])
        ex_zh = clean_val(r[hmap.get('中文例句 (Example ZH-CN)', 13)])
        ex_fr = clean_val(r[hmap.get('Phrase française (Example FR)', 14)])
        ex_ja = clean_val(r[hmap.get('日本語の例文 (Example JA)', 15)])
        ex_hi = clean_val(r[hmap.get('हिन्दी उदाहरण (Example HI)', 16)])
        ex_vi = clean_val(r[hmap.get('Câu ví dụ tiếng Việt (Example VI)', 17)])

        item = {
            'id': db_id,
            'word': word,
            'phonics': phonics,
            'meaning': meaning,
            'meaning_zh': meaning_zh,
            'meaning_fr': meaning_fr,
            'meaning_ja': meaning_ja,
            'meaning_hi': meaning_hi,
            'meaning_vi': meaning_vi,
            'category': category,
            'grade_level': '중등단어',
            'grade_level_ko': '중등단어',
            'grade_level_zh': '初中英语',
            'grade_level_fr': 'Anglais Collège',
            'grade_level_ja': '中学生英語',
            'grade_level_vi': 'Tiếng Anh THCS',
            'grade_level_hi': 'माध्यमिक अंग्रेजी',
            'example_en': ex_en,
            'example_ko': ex_ko,
            'example_zh': ex_zh,
            'example_fr': ex_fr,
            'example_ja': ex_ja,
            'example_hi': ex_hi,
            'example_vi': ex_vi,
        }
        items.append(item)
    print(f"Parsed {len(items)} middle school words.")
    return items

def parse_high():
    fp = r'D:\Yc_Back_Live\영창정밀_노트북\english_edu\word\multilingual\highschool_suneung_vocab_3000_multilingual.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    hmap = {h: idx for idx, h in enumerate(headers)}

    items = []
    for r in rows[1:]:
        num = r[hmap.get('번호', 0)]
        if not num: continue
        db_id = int(num) + 3000 # 1~3000 ➔ 3001~6000
        word = clean_val(r[hmap.get('영어 단어 (Word)', 1)])
        phonics = clean_val(r[hmap.get('발음기호 (IPA)', 2)])
        meaning = clean_val(r[hmap.get('한국어 뜻 (Meaning)', 3)])
        meaning_zh = clean_val(r[hmap.get('中文释义 (Meaning ZH-CN)', 4)])
        meaning_fr = clean_val(r[hmap.get('Français (Meaning FR)', 5)])
        meaning_ja = clean_val(r[hmap.get('日本語の意味 (Meaning JA)', 6)])
        meaning_hi = clean_val(r[hmap.get('हिन्दी अर्थ (Meaning HI)', 7)])
        meaning_vi = clean_val(r[hmap.get('Nghĩa tiếng Việt (Meaning VI)', 8)])
        category_raw = clean_val(r[hmap.get('주제 (Category)', 9)])
        category = f"고등 - {category_raw}" if category_raw and not category_raw.startswith('고등') else category_raw
        ex_en = clean_val(r[hmap.get('영어 예문 (Example EN)', 13)])
        ex_ko = clean_val(r[hmap.get('한국어 해석 (Example KO)', 14)])
        ex_zh = clean_val(r[hmap.get('中文例句 (Example ZH-CN)', 15)])
        ex_fr = clean_val(r[hmap.get('Phrase française (Example FR)', 16)])
        ex_ja = clean_val(r[hmap.get('日本語の例文 (Example JA)', 17)])
        ex_hi = clean_val(r[hmap.get('हिन्दी उदाहरण (Example HI)', 18)])
        ex_vi = clean_val(r[hmap.get('Câu ví dụ tiếng Việt (Example VI)', 19)])

        item = {
            'id': db_id,
            'word': word,
            'phonics': phonics,
            'meaning': meaning,
            'meaning_zh': meaning_zh,
            'meaning_fr': meaning_fr,
            'meaning_ja': meaning_ja,
            'meaning_hi': meaning_hi,
            'meaning_vi': meaning_vi,
            'category': category,
            'grade_level': '고등단어',
            'grade_level_ko': '고등단어',
            'grade_level_zh': '高中/高考英语',
            'grade_level_fr': 'Anglais Lycée',
            'grade_level_ja': '高校生英語',
            'grade_level_vi': 'Tiếng Anh THPT',
            'grade_level_hi': 'उच्च माध्यमिक अंग्रेजी',
            'example_en': ex_en,
            'example_ko': ex_ko,
            'example_zh': ex_zh,
            'example_fr': ex_fr,
            'example_ja': ex_ja,
            'example_hi': ex_hi,
            'example_vi': ex_vi,
        }
        items.append(item)
    print(f"Parsed {len(items)} high school words.")
    return items

def update_single_word_in_db(item):
    db_id = item['id']
    payload = {k: v for k, v in item.items() if k != 'id'}
    url = f"{SUPABASE_URL}/rest/v1/words?id=eq.{db_id}"

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal'
        },
        method='PATCH'
    )

    for retry in range(3):
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                return (db_id, True, None)
        except Exception as e:
            if retry == 2:
                return (db_id, False, str(e))
            time.sleep(0.3 * (retry + 1))
    return (db_id, False, 'Max retries exceeded')

def main():
    print("==================================================")
    print("🚀 다국어 엑셀 3종 전체 5,000개 단어 Supabase DB 업로드 시작")
    print("==================================================")

    elem_words = parse_elementary()
    middle_words = parse_middle()
    high_words = parse_high()

    all_words = elem_words + middle_words + high_words
    total = len(all_words)
    print(f"\n총 업로드 대상 단어: {total}개 (초등 {len(elem_words)} + 중등 {len(middle_words)} + 고등 {len(high_words)})")

    success_count = 0
    fail_count = 0
    start_time = time.time()

    print("\nSupabase 클라우드 DB 비동기 고속 PATCH 업데이트 진행 중...")
    with ThreadPoolExecutor(max_workers=30) as executor:
        futures = {executor.submit(update_single_word_in_db, item): item['id'] for item in all_words}
        completed = 0
        for future in as_completed(futures):
            completed += 1
            db_id, ok, err = future.result()
            if ok:
                success_count += 1
            else:
                fail_count += 1
                print(f"❌ ID {db_id} 업데이트 실패: {err}")

            if completed % 500 == 0 or completed == total:
                elapsed = time.time() - start_time
                print(f"진행률: {completed}/{total} ({completed/total*100:.1f}%) - 소요시간: {elapsed:.1f}초 (성공: {success_count}, 실패: {fail_count})")

    elapsed_total = time.time() - start_time
    print(f"\n🎉 5,000개 전체 단어 DB 업로드 완료! (성공: {success_count}, 실패: {fail_count}, 총 소요시간: {elapsed_total:.1f}초)")

if __name__ == '__main__':
    main()
