import openpyxl, glob, json, sys, io, os
import urllib.request

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SUPABASE_URL = 'https://sqonhhqosyszncjfoxfd.supabase.co'
SUPABASE_KEY = 'sb_publishable_1trPlZQEdVKMvUYQNV5aVA_nSQqOiuo'

def fetch_all_db_words():
    all_rows = []
    offset = 0
    limit = 1000
    while True:
        req = urllib.request.Request(
            f'{SUPABASE_URL}/rest/v1/words?select=id,word,meaning,grade_level&offset={offset}&limit={limit}',
            headers={
                'apikey': SUPABASE_KEY,
                'Authorization': f'Bearer {SUPABASE_KEY}'
            }
        )
        try:
            with urllib.request.urlopen(req) as resp:
                data = json.loads(resp.read().decode('utf-8'))
                if not data:
                    break
                all_rows.extend(data)
                if len(data) < limit:
                    break
                offset += limit
        except Exception as e:
            print('Fetch DB error:', e)
            break
    return all_rows

def main():
    print('🚀 중국어 데이터 엑셀 파싱 및 데이터베이스/로컬 JSON 통합 동기화 작업 시작...\n')

    # 1. 엑셀 파싱 (초등, 중등, 고등)
    parsed_items = {}  # (grade, word_lower) -> { meaning_zh, example_zh }

    # 초등
    elem_path = 'word/elementary_words_ko_zh.xlsx'
    if os.path.exists(elem_path):
        wb = openpyxl.load_workbook(elem_path, data_only=True)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        print(f'📖 초등 엑셀 읽기 완료 ({ws.max_row-1}행)...')
        for r in range(2, ws.max_row+1):
            w = str(ws.cell(r, 2).value or '').strip()
            zh_m = str(ws.cell(r, 5).value or '').strip()
            zh_ex = str(ws.cell(r, 11).value or '').strip()
            if w and zh_m and zh_m != 'None':
                parsed_items[('초등단어', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}
                parsed_items[('general', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}

    # 중등
    mid_path = 'word/middle_school_words_ko_zh.xlsx'
    if os.path.exists(mid_path):
        wb = openpyxl.load_workbook(mid_path, data_only=True)
        ws = wb['전체 목록'] if '전체 목록' in wb.sheetnames else wb.active
        print(f'📖 중등 엑셀 읽기 완료 ({ws.max_row-1}행)...')
        for r in range(2, ws.max_row+1):
            w = str(ws.cell(r, 2).value or '').strip()
            zh_m = str(ws.cell(r, 6).value or '').strip()
            zh_ex = str(ws.cell(r, 10).value or '').strip()
            if w and zh_m and zh_m != 'None':
                parsed_items[('중등단어', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}
                if ('general', w.lower()) not in parsed_items:
                    parsed_items[('general', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}

    # 고등
    high_path = 'word/highschool_suneung_vocab_3000_ko_zh.xlsx'
    if os.path.exists(high_path):
        wb = openpyxl.load_workbook(high_path, data_only=True)
        ws = wb['vocabulary'] if 'vocabulary' in wb.sheetnames else wb.active
        print(f'📖 고등 엑셀 읽기 완료 ({ws.max_row-1}행)...')
        for r in range(2, ws.max_row+1):
            w = str(ws.cell(r, 2).value or '').strip()
            zh_m = str(ws.cell(r, 5).value or '').strip()
            zh_ex = str(ws.cell(r, 11).value or '').strip()
            if w and zh_m and zh_m != 'None':
                parsed_items[('고등/수능', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}
                parsed_items[('고등단어', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}
                if ('general', w.lower()) not in parsed_items:
                    parsed_items[('general', w.lower())] = {'meaning_zh': zh_m, 'example_zh': zh_ex if zh_ex != 'None' else ''}

    print(f'✅ 총 {len(parsed_items)}개의 중국어 어휘 매핑 생성 완료!\n')

    # 2. DB 단어 매칭 및 업데이트 페이로드 생성
    db_words = fetch_all_db_words()
    print(f'📊 Supabase DB에서 총 {len(db_words)}개 단어 로드 완료.')

    updates_payload = []
    matched_count = 0

    for item in db_words:
        w_id = item['id']
        w_str = (item.get('word') or '').strip().lower()
        grade = (item.get('grade_level') or '').strip()

        zh_data = parsed_items.get((grade, w_str)) or parsed_items.get(('general', w_str))
        if zh_data:
            matched_count += 1
            updates_payload.append({
                'id': w_id,
                'word': item.get('word'),
                'meaning_zh': zh_data['meaning_zh'],
                'example_zh': zh_data['example_zh']
            })

    print(f'🎯 Supabase DB 단어 중 {matched_count}개 중국어 매칭 성공!')

    # 3. REST API Upsert 시도
    if updates_payload:
        chunk_size = 100
        success_db = 0
        column_exists = True

        for i in range(0, len(updates_payload), chunk_size):
            chunk = updates_payload[i:i+chunk_size]
            payload_bytes = json.dumps(chunk, ensure_ascii=False).encode('utf-8')

            req = urllib.request.Request(
                f'{SUPABASE_URL}/rest/v1/words',
                data=payload_bytes,
                headers={
                    'apikey': SUPABASE_KEY,
                    'Authorization': f'Bearer {SUPABASE_KEY}',
                    'Content-Type': 'application/json',
                    'Prefer': 'resolution=merge-duplicates'
                },
                method='POST'
            )
            try:
                with urllib.request.urlopen(req) as resp:
                    if resp.status in (200, 201, 204):
                        success_db += len(chunk)
            except urllib.error.HTTPError as e:
                err_body = e.read().decode('utf-8')
                print(f'⚠️ DB 업로드 응답 ({e.code}):', err_body)
                if 'meaning_zh' in err_body or 'PGRST204' in err_body:
                    column_exists = False
                    print('📌 Supabase DB에 meaning_zh, example_zh 컬럼 추가 SQL 실행이 필요합니다.')
                    break

        if column_exists:
            print(f'🎉 Supabase DB에 {success_db}개 중국어 뜻/예문 저장 성공!')

    # 4. 로컬 JSON 파일들 동기화 업데이트 (data/ 및 next_app/data/)
    local_files = [
        'data/parsed_800_words.json',
        'data/middle_school_words.json',
        'data/highschool_parsed.json'
    ]

    for lf in local_files:
        if os.path.exists(lf):
            with open(lf, 'r', encoding='utf-8-sig') as f:
                content = json.load(f)

            mod_count = 0
            for row in content:
                w_str = (row.get('word') or '').strip().lower()
                zh_info = parsed_items.get(('general', w_str))
                if zh_info:
                    row['meaning_zh'] = zh_info['meaning_zh']
                    row['meaningZh'] = zh_info['meaning_zh']
                    row['example_zh'] = zh_info['example_zh']
                    row['exampleZh'] = zh_info['example_zh']
                    mod_count += 1

            with open(lf, 'w', encoding='utf-8') as f:
                json.dump(content, f, ensure_ascii=False, indent=2)

            print(f'📝 {lf} 데이터 파일 {mod_count}개 중국어 어휘 업데이트 완료!')

if __name__ == '__main__':
    main()
