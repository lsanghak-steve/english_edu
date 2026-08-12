
import openpyxl, json, sys, os

sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r"D:\Yc_Back_Live\영창정밀_노트북\english_edu\word\highschool_word.xlsx")
sheet = wb['vocabulary']

words = []
for r in range(2, sheet.max_row + 1):
    num = sheet.cell(row=r, column=1).value
    word = sheet.cell(row=r, column=2).value
    ipa = sheet.cell(row=r, column=3).value
    meaning = sheet.cell(row=r, column=4).value
    category = sheet.cell(row=r, column=5).value
    level = sheet.cell(row=r, column=6).value
    status = sheet.cell(row=r, column=7).value
    file_name = sheet.cell(row=r, column=8).value
    example_en = sheet.cell(row=r, column=9).value
    example_ko = sheet.cell(row=r, column=10).value

    if word:
        word_str = str(word).strip()
        words.append({
            'excel_no': num,
            'word': word_str,
            'phonics': str(ipa or '').strip(),
            'meaning': str(meaning or '').strip(),
            'category': f"고등 - {str(category or '일반').strip()}",
            'grade_level': '고등단어',
            'example_en': str(example_en or '').strip(),
            'example_ko': str(example_ko or '').strip(),
            'image_url': f"/word_img/{word_str.lower()}.png"
        })

out_path = r"D:\\Yc_Back_Live\\영창정밀_노트북\\english_edu\\highschool_parsed.json"
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(words, f, ensure_ascii=False, indent=2)

print(f'Successfully exported {len(words)} high school words to JSON.')
