import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def update_excel_files_lowercase():
    base_dir = os.path.dirname(__file__)
    files = [
        os.path.join(base_dir, 'word', 'elementary_words_ko_zh.xlsx'),
        os.path.join(base_dir, 'word', 'middle_school_words_ko_zh.xlsx'),
        os.path.join(base_dir, 'word', 'highschool_suneung_vocab_3000_ko_zh.xlsx')
    ]

    print("====================================================")
    print("📖 [엑셀 단어장 단어 컬럼 소문자 일괄 변환 시작]")
    print("====================================================\n")

    for fpath in files:
        if not os.path.exists(fpath):
            continue

        fname = os.path.basename(fpath)
        wb = openpyxl.load_workbook(fpath)
        ws = wb.active

        headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
        word_col_idx = 2  # 기본 B열

        for idx, h in enumerate(headers, 1):
            if h in ['word', '단어', 'english']:
                word_col_idx = idx
                break

        count = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=word_col_idx)
            val = str(cell.value or '').strip()
            if val and val != val.lower():
                cell.value = val.lower()
                count += 1

        wb.save(fpath)
        print(f"  ✅ [{fname}] 총 {count}개 단어 소문자 변환 저장 완료!")

    print("\n🎉 모든 엑셀 단어장의 단어가 소문자로 완벽히 변환되었습니다!\n")

if __name__ == '__main__':
    update_excel_files_lowercase()
