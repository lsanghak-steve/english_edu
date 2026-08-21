import os
import sys
import re
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def get_local_image_set():
    img_dir = os.path.join(os.path.dirname(__file__), 'next_app', 'public', 'word_img')
    normalized_set = set()
    total_files = 0

    if os.path.exists(img_dir):
        for f in os.listdir(img_dir):
            if f.endswith('.png') and not f.startswith('_'):
                total_files += 1
                base = f.replace('.png', '').strip().lower()
                normalized_set.add(base)
                normalized_set.add(re.sub(r'[\s\-_]', '', base))
    
    return normalized_set, total_files

def analyze_excel(file_path, grade_name, local_images):
    if not os.path.exists(file_path):
        return {'grade': grade_name, 'total': 0, 'exist': 0, 'missing': []}

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # 헤더 행 찾기
    headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
    word_idx = -1
    meaning_idx = -1

    for idx, h in enumerate(headers):
        if h in ['word', '단어', 'english']:
            word_idx = idx
        elif h in ['meaning', '뜻', 'korean', '의미']:
            meaning_idx = idx

    if word_idx == -1:
        word_idx = 1
    if meaning_idx == -1:
        meaning_idx = 3

    missing_list = []
    exist_count = 0
    total_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) <= word_idx or not row[word_idx]:
            continue
        
        raw_word = str(row[word_idx]).strip()
        raw_meaning = str(row[meaning_idx] if len(row) > meaning_idx and row[meaning_idx] else '').strip()

        if not raw_word or raw_word.lower() == 'word':
            continue

        total_count += 1
        clean_word = raw_word.lower()
        clean_no_space = re.sub(r'[\s\-_]', '', clean_word)

        if clean_word in local_images or clean_no_space in local_images:
            exist_count += 1
        else:
            missing_list.append({
                'row': row_idx,
                'word': raw_word,
                'meaning': raw_meaning,
                'grade': grade_name
            })

    return {
        'grade': grade_name,
        'file': os.path.basename(file_path),
        'total': total_count,
        'exist': exist_count,
        'missing': missing_list
    }

def main():
    print("====================================================")
    print("🔍 [초/중/고 5,000개 마스터 단어장 이미지 전수 검사]")
    print("====================================================\n")

    local_images, total_img_files = get_local_image_set()
    print(f"📁 next_app/public/word_img 내 보유 이미지: 총 {total_img_files}개\n")

    base_dir = os.path.dirname(__file__)
    targets = [
        ('초등단어 (800)', os.path.join(base_dir, 'word', 'elementary_words_ko_zh.xlsx')),
        ('중등단어 (1,200)', os.path.join(base_dir, 'word', 'middle_school_words_ko_zh.xlsx')),
        ('고등/수능단어 (3,000)', os.path.join(base_dir, 'word', 'highschool_suneung_vocab_3000_ko_zh.xlsx'))
    ]

    all_results = []
    grand_total = 0
    grand_exist = 0
    grand_missing = []

    for grade_name, fpath in targets:
        res = analyze_excel(fpath, grade_name, local_images)
        all_results.append(res)
        grand_total += res['total']
        grand_exist += res['exist']
        grand_missing.extend(res['missing'])

        rate = (res['exist'] / res['total'] * 100) if res['total'] > 0 else 0
        print(f"📖 [{grade_name}] (파일명: {res.get('file', '')})")
        print(f"  - 총 단어 수: {res['total']}개")
        print(f"  - 🖼️ 이미지 보유: {res['exist']}개 ({rate:.1f}%)")
        print(f"  - ❌ 미보유 단어: {len(res['missing'])}개\n")

    grand_rate = (grand_exist / grand_total * 100) if grand_total > 0 else 0
    print("====================================================")
    print(f"🏆 [전체 마스터 5,000단어 종합]")
    print(f"  - 총 단어 수: {grand_total}개")
    print(f"  - 🖼️ 전체 보유 이미지: {grand_exist}개 ({grand_rate:.1f}%)")
    print(f"  - ❌ 전체 미보유 단어: {len(grand_missing)}개 ({100 - grand_rate:.1f}%)")
    print("====================================================\n")

    # 결과 JSON 저장
    out_json = os.path.join(base_dir, 'missing_image_words_full_report.json')
    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump({
            'grand_total': grand_total,
            'grand_exist': grand_exist,
            'grand_missing_count': len(grand_missing),
            'coverage_rate': f"{grand_rate:.1f}%",
            'results_by_grade': all_results,
            'all_missing_words': grand_missing
        }, f, ensure_ascii=False, indent=2)

    print(f"💾 전체 미보유 단어 상세 리포트 JSON 저장 완료: {out_json}\n")

    # 카테고리별 미보유 단어 목록 출력
    for res in all_results:
        print(f"--- ❌ [{res['grade']}] 미보유 단어 목록 (총 {len(res['missing'])}개) ---")
        if len(res['missing']) == 0:
            print("  🎉 미보유 단어 없음 (100% 이미지 보유 완료!)")
        else:
            for idx, m in enumerate(res['missing'], 1):
                print(f"  [{idx:3d}] {m['word']:<20} | 뜻: {m['meaning']}")
        print()

if __name__ == '__main__':
    main()
