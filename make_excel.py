import csv
import sys

# 1. 초등 3~5학년 필수 권장 기초 영단어 100선 데이터
words_data = [
    # 과일 & 음식 (1~10)
    ["Apple", "[애플]", "사과", "과일/음식 🍎", "I eat an apple.", "나는 사과를 먹어요."],
    ["Banana", "[버내너]", "바나나", "과일/음식 🍌", "Monkeys like bananas.", "원숭이는 바나나를 좋아해요."],
    ["Milk", "[밀크]", "우유", "과일/음식 🥛", "Drink warm milk.", "따뜻한 우유를 마셔요."],
    ["Water", "[워터]", "물", "과일/음식 💧", "I drink water.", "나는 물을 마셔요."],
    ["Bread", "[브레드]", "빵", "과일/음식 🍞", "I eat bread for breakfast.", "나는 아침으로 빵을 먹어요."],
    ["Egg", "[에그]", "계란/달걀", "과일/음식 🥚", "I like fried eggs.", "나는 계란 프라이를 좋아해요."],
    ["Juice", "[주스]", "주스", "과일/음식 🧃", "Orange juice is sweet.", "오렌지 주스는 달콤해요."],
    ["Rice", "[라이스]", "밥/쌀", "과일/음식 🍚", "We eat rice every day.", "우리는 매일 밥을 먹어요."],
    ["Cake", "[케이크]", "케이크", "과일/음식 🎂", "Happy birthday cake!", "생일 축하 케이크!"],
    ["Candy", "[캔디]", "사탕", "과일/음식 🍬", "Sweet candy is yummy.", "달콤한 사탕은 맛있어요."],

    # 동물 (11~20)
    ["Dog", "[독]", "개/강아지", "동물 🐶", "My dog wags its tail.", "내 강아지가 꼬리를 쳐요."],
    ["Cat", "[캣]", "고양이", "동물 🐱", "The cat sleeps on the bed.", "고양이가 침대에서 자요."],
    ["Bird", "[버드]", "새", "동물 🐦", "The bird sings softly.", "새가 부드럽게 노래해요."],
    ["Fish", "[피쉬]", "물고기", "동물 🐟", "Fish swim in the river.", "물고기가 강에서 헤엄쳐요."],
    ["Duck", "[덕]", "오리", "동물 🦆", "Ducks swim well.", "오리는 수영을 잘해요."],
    ["Pig", "[피그]", "돼지", "동물 🐷", "Pigs are cute.", "돼지는 귀여워요."],
    ["Bear", "[베어]", "곰", "동물 🐻", "Bears love honey.", "곰은 꿀을 좋아해요."],
    ["Rabbit", "[래빗]", "토끼", "동물 🐰", "Rabbits jump high.", "토끼는 높이 뛰어올라요."],
    ["Lion", "[라이언]", "사자", "동물 🦁", "Lions are strong.", "사자는 힘이 세요."],
    ["Monkey", "[멍키]", "원숭이", "동물 🐒", "Monkeys climb trees.", "원숭이는 나무를 타요."],

    # 가족 & 사람 (21~30)
    ["Father", "[파더]", "아버지/아빠", "가족/사람 👨", "My father is kind.", "우리 아버지는 친절해요."],
    ["Mother", "[마더]", "어머니/엄마", "가족/사람 👩", "I love my mother.", "나는 엄마를 사랑해요."],
    ["Brother", "[브라더]", "남동생/형/오빠", "가족/사람 👦", "My brother plays soccer.", "내 남동생은 축구를 해요."],
    ["Sister", "[시스터]", "여동생/누나/언니", "가족/사람 👧", "My sister plays the piano.", "내 여동생은 피아노를 쳐요."],
    ["Friend", "[프렌드]", "친구", "가족/사람 🧑‍🤝‍🧑", "We are good friends.", "우리는 좋은 친구예요."],
    ["Teacher", "[티처]", "선생님", "가족/사람 👩‍🏫", "The teacher is smiling.", "선생님이 미소 짓고 계셔요."],
    ["Baby", "[베이비]", "아기", "가족/사람 👶", "The baby is sleeping.", "아기가 자고 있어요."],
    ["Doctor", "[닥터]", "의사", "가족/사람 🧑‍⚕️", "The doctor helps sick people.", "의사는 아픈 사람을 도와요."],
    ["Boy", "[보이]", "소년/남자아이", "가족/사람 👦", "The boy runs fast.", "그 소년은 빠르게 달려요."],
    ["Girl", "[걸]", "소녀/여자아이", "가족/사람 👧", "The girl dances nicely.", "그 소녀는 춤을 잘 춰요."],

    # 학교 & 학용품 (31~40)
    ["School", "[스쿨]", "학교", "학교/학용품 🏫", "I go to school.", "나는 학교에 가요."],
    ["Book", "[북]", "책", "학교/학용품 📖", "Read a good book.", "좋은 책을 읽어요."],
    ["Pencil", "[펜슬]", "연필", "학교/학용품 ✏️", "I write with a pencil.", "나는 연필로 글을 써요."],
    ["Desk", "[데스크]", "책상", "학교/학용품 🪑", "Clean your desk.", "책상을 깨끗이 정리해요."],
    ["Chair", "[체어]", "의자", "학교/학용품 🪑", "Sit on the chair.", "의자에 앉으세요."],
    ["Bag", "[백]", "가방", "학교/학용품 🎒", "My bag is blue.", "내 가방은 파란색이에요."],
    ["Eraser", "[이레이저]", "지우개", "학교/학용품 🧹", "Pass me the eraser.", "지우개 좀 전해줘."],
    ["Ruler", "[룰러]", "자", "학교/학용품 📏", "Use a ruler to draw lines.", "선을 그릴 때 자를 사용해요."],
    ["Class", "[클래스]", "반/수업", "학교/학용품 🏫", "English class is fun.", "영어 수업은 재미있어요."],
    ["Computer", "[컴퓨터]", "컴퓨터", "학교/학용품 💻", "I study with a computer.", "나는 컴퓨터로 공부해요."],

    # 몸 & 얼굴 (41~50)
    ["Eye", "[아이]", "눈", "몸/얼굴 👀", "Close your eyes.", "눈을 감으세요."],
    ["Ear", "[이어]", "귀", "몸/얼굴 👂", "I listen with my ears.", "나는 귀로 들어요."],
    ["Mouth", "[마우스]", "입", "몸/얼굴 👄", "Open your mouth.", "입을 벌리세요."],
    ["Nose", "[노우즈]", "코", "몸/얼굴 👃", "Touch your nose.", "코를 터치하세요."],
    ["Hand", "[핸드]", "손", "몸/얼굴 ✋", "Wash your hands.", "손을 씻으세요."],
    ["Foot", "[풋]", "발", "몸/얼굴 🦶", "My foot is warm.", "내 발은 따뜻해요."],
    ["Head", "[헤드]", "머리", "몸/얼굴 🗣️", "Nod your head.", "머리를 끄덕이세요."],
    ["Face", "[페이스]", "얼굴", "몸/얼굴 😊", "Smile with a happy face.", "행복한 얼굴로 웃어요."],
    ["Arm", "[암]", "팔", "몸/얼굴 💪", "Raise your arms.", "팔을 들어 올리세요."],
    ["Leg", "[렉]", "다리", "몸/얼굴 🦵", "My legs are long.", "내 다리는 길어요."],

    # 숫자 & 색상 (51~60)
    ["One", "[원]", "숫자 1", "숫자/색상 1️⃣", "I have one apple.", "나는 사과 1개가 있어요."],
    ["Two", "[투]", "숫자 2", "숫자/색상 2️⃣", "Two eyes see well.", "두 눈은 잘 봐요."],
    ["Three", "[쓰리]", "숫자 3", "숫자/색상 3️⃣", "Three little pigs.", "아기 돼지 삼 형제."],
    ["Red", "[레드]", "빨간색", "숫자/색상 🔴", "Red is my favorite color.", "빨간색은 내가 제일 좋아하는 색이에요."],
    ["Blue", "[블루]", "파란색", "숫자/색상 🔵", "The sky is blue.", "하늘이 파란색이에요."],
    ["Yellow", "[옐로우]", "노란색", "숫자/색상 🟡", "Sunflowers are yellow.", "해바라기는 노란색이에요."],
    ["Green", "[그린]", "초록색", "숫자/색상 🟢", "Grass is green.", "잔디는 초록색이에요."],
    ["White", "[화이트]", "하얀색", "숫자/색상 ⚪", "Snow is white.", "눈은 하얀색이에요."],
    ["Black", "[블랙]", "검은색", "숫자/색상 ⬛", "I like black shoes.", "나는 검은 신발을 좋아해요."],
    ["Pink", "[핑크]", "분홍색", "숫자/색상 🩷", "Pink flowers are pretty.", "분홍 꽃들이 예뻐요."],

    # 날씨 & 자연 (61~70)
    ["Sun", "[썬]", "태양/해", "날씨/자연 ☀️", "The sun is bright.", "태양이 밝게 빛나야."],
    ["Moon", "[문]", "달", "날씨/자연 🌙", "The moon is round.", "달이 둥글어요."],
    ["Star", "[스타]", "별", "날씨/자연 ⭐", "Stars shine at night.", "밤에 별들이 빛나요."],
    ["Sky", "[스카이]", "하늘", "날씨/자연 ☁️", "Look at the high sky.", "높은 하늘을 보세요."],
    ["Rain", "[레인]", "비", "날씨/자연 🌧️", "Rain falls from the sky.", "하늘에서 비가 내려요."],
    ["Snow", "[스노우]", "눈(자연)", "날씨/자연 ❄️", "I like white snow.", "나는 하얀 눈을 좋아해요."],
    ["Tree", "[트리]", "나무", "날씨/자연 🌳", "Birds live in the tree.", "새들이 나무에 살아요."],
    ["Flower", "[플라워]", "꽃", "날씨/자연 🌸", "Smell the sweet flower.", "달콤한 꽃향기를 맡아보세요."],
    ["Sea", "[시]", "바다", "날씨/자연 🌊", "The sea is vast.", "바다는 아주 넓어요."],
    ["Wind", "[윈드]", "바람", "날씨/자연 🌬️", "Cool wind is blowing.", "시원한 바람이 불어요."],

    # 행동 & 동사 (71~80)
    ["Go", "[고우]", "가다", "행동/동사 🏃", "Go to school.", "학교에 가요."],
    ["Come", "[컴]", "오다", "행동/동사 🚶", "Come here, please.", "이리로 오세요."],
    ["Eat", "[잇]", "먹다", "행동/동사 🍽️", "Eat healthy food.", "건강한 음식을 먹어요."],
    ["Drink", "[드링크]", "마시다", "행동/동사 🥤", "Drink fresh water.", "신선한 물을 마셔요."],
    ["Sleep", "[슬립]", "자다", "행동/동사 😴", "Sleep well at night.", "밤에 잘 자요."],
    ["Run", "[런]", "달리다", "행동/동사 🏃", "Run fast in the park.", "공원에서 빠르게 달려요."],
    ["Walk", "[워크]", "걷다", "행동/동사 🚶", "Walk slowly.", "천천히 걸어요."],
    ["Sing", "[싱]", "노래하다", "행동/동사 🎤", "Sing a song happily.", "즐겁게 노래를 불러요."],
    ["Dance", "[댄스]", "춤추다", "행동/동사 💃", "Dance to the music.", "음악에 맞춰 춤춰요."],
    ["Read", "[리드]", "읽다", "행동/동사 📖", "Read an interesting story.", "재미있는 이야기를 읽어요."],

    # 상태 & 형용사 (81~90)
    ["Happy", "[해피]", "행복한/기쁜", "상태/형용사 😊", "I am very happy today.", "나는 오늘 아주 행복해요."],
    ["Sad", "[새드]", "슬픈", "상태/형용사 😢", "Don't be sad.", "슬퍼하지 마세요."],
    ["Big", "[빅]", "크기가 큰", "상태/형용사 🐘", "That is a big elephant.", "저것은 큰 코끼리예요."],
    ["Small", "[스몰]", "크기가 작은", "상태/형용사 🐭", "The mouse is small.", "생쥐는 작아요."],
    ["Good", "[굿]", "좋은/훌륭한", "상태/형용사 👍", "Have a good day!", "좋은 하루 보내세요!"],
    ["Hot", "[핫]", "뜨거운/따뜻한", "상태/형용사 ♨️", "Hot tea is warm.", "뜨거운 차는 따뜻해요."],
    ["Cold", "[콜드]", "차가운/추운", "상태/형용사 🧊", "Cold ice cream.", "차가운 아이스크림."],
    ["Fast", "[패스트]", "빠른", "상태/형용사 ⚡", "The cheetah is fast.", "치타는 빨라요."],
    ["Cute", "[큐트]", "귀여운", "상태/형용사 🐱", "The puppy is cute.", "강아지가 귀여워요."],
    ["Pretty", "[프리티]", "예쁜", "상태/형용사 🌸", "Pretty pink dress.", "예쁜 분홍색 드레스."],

    # 장소 & 사물 (91~100)
    ["Home", "[홈]", "집/우리집", "장소/사물 🏠", "Welcome home!", "집에 온 걸 환영해!"],
    ["Park", "[파크]", "공원", "장소/사물 🏞️", "Play in the park.", "공원에서 놀아요."],
    ["Room", "[룸]", "방", "장소/사물 🚪", "My room is clean.", "내 방은 깨끗해요."],
    ["Car", "[카]", "자동차", "장소/사물 🚗", "The red car moves fast.", "빨간 자동차가 빠르게 움직여요."],
    ["Bus", "[버스]", "버스", "장소/사물 🚌", "Take the school bus.", "스쿨버스를 타요."],
    ["Door", "[도어]", "문", "장소/사물 🚪", "Open the door.", "문을 열어주세요."],
    ["Window", "[윈도우]", "창문", "장소/사물 🪟", "Look out the window.", "창밖을 보세요."],
    ["Clock", "[클락]", "시계", "장소/사물 ⏰", "Check the clock.", "시계를 확인해보세요."],
    ["Toy", "[토이]", "장난감", "장소/사물 🧸", "I play with my toy.", "나는 장난감을 가지고 놀아요."],
    ["Game", "[게임]", "게임/놀이", "장소/사물 🎮", "Playing games is fun.", "게임하는 것은 재미있어요."]
]

headers = ["번호", "영어 단어 (Word)", "한글 발음 (Phonics)", "한국어 뜻 (Meaning)", "주제 (Category)", "영어 예문 (Example EN)", "한국어 해석 (Example KO)"]

# 1. UTF-8 CSV 파일 작성 (엑셀에서 바로 열림)
csv_filename = "elementary_words.csv"
with open(csv_filename, "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(headers)
    for idx, row in enumerate(words_data, start=1):
        writer.writerow([idx] + row)

print(f"✅ CSV 파일 생성 완료: {csv_filename}")

# 2. openpyxl 라이브러리를 통한 진짜 .xlsx 엑셀 파일 생성 시도
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "초등 필수 영단어 100선"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="맑은 고딕", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 헤더 쓰기
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # 데이터 쓰기
    for idx, row_data in enumerate(words_data, start=1):
        full_row = [idx] + row_data
        ws.append(full_row)
        row_num = idx + 1
        for col_num in range(1, len(full_row) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            if col_num in [1, 3, 5]: # 번호, 한글발음, 카테고리 중앙 정렬
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 열 너비 자동 조절
    column_widths = [8, 18, 16, 18, 18, 35, 35]
    for i, width in enumerate(column_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    xlsx_filename = "elementary_words.xlsx"
    wb.save(xlsx_filename)
    print(f"✅ XLSX 엑셀 파일 생성 완료: {xlsx_filename}")

except ImportError:
    print("openpyxl 패키지가 없습니다. CSV 파일만 우선 작성했습니다.")
